# ting som skal gjøres: Lag en variabel med siste celle. så skriv siste celle + antall celler
# siden antall sikkerhetsvakter varierer

from datetime import date, timedelta, datetime
import calendar
import locale
locale.setlocale(locale.LC_ALL, 'no_NO.utf8') #trengte ikke å gjøre det på jobb-pc

import pandas as pd
from dateutil.easter import *
from openpyxl import Workbook, load_workbook, worksheet
from openpyxl.cell import cell
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, colors
from openpyxl.styles.borders import Border, Side, BORDER_THIN

# Velg hvilket år det skal lages turnus for
år = int(input("Tast inn år:  "))

# Sjekk om det er stortingsvalg
valgår = pd.date_range(start='10/13/2021', periods=20, freq='4Y').year

if år in valgår:
    print('valg')
else:
    print('ikke valg')

#sjekk om det er skuddår
skuddår = calendar.isleap(år)

if skuddår == True:
    februar_skuddår = 29
    ekstra_dager = 2
else:
    februar_skuddår = 28
    ekstra_dager = 3
    
print(februar_skuddår)

# Last turnus-Excelfil
wb_turnus = load_workbook('Turnus_blank.xlsx')
wb_turnusmal = load_workbook('turnus_mal.xlsx')

# Velg worksheet for turnusmal
ws_turnus_mal = wb_turnusmal['Helturnus']

# Velg worksheet for turnus
ws_turnus_jan = wb_turnus['Januar']
ws_turnus_feb = wb_turnus['Februar']
ws_turnus_mar = wb_turnus['Mars']
ws_turnus_apr = wb_turnus['April']
ws_turnus_mai = wb_turnus['Mai']
ws_turnus_jun = wb_turnus['Juni']
ws_turnus_jul = wb_turnus['Juli']
ws_turnus_aug = wb_turnus['August']
ws_turnus_sep = wb_turnus['September']
ws_turnus_okt = wb_turnus['Oktober']
ws_turnus_nov = wb_turnus['November']
ws_turnus_des = wb_turnus['Desember']

# Definer bevegelige helligdager 
#år = datetime.today().year
første_påskedag = easter(år) # 1. påskedag kan falle mellom 22. mars og 25. april
skjærtorsdag = første_påskedag - timedelta(days=3)
langfredag = første_påskedag - timedelta(days=2)
påskeaften = første_påskedag - timedelta(days=1)
andre_påskedag = første_påskedag + timedelta(days=1)
himmelspretten = første_påskedag + timedelta(days=39)#Kristi himmelfartsdag faller dermed som regel i mai, men den kan også falle på en av de tre første dagene i juni. Dette skjedde sist i 1943, 2000 og 2011, og skjer neste gang i 2038. Den kan også falle på den siste dagen i april, men dét skjedde sist i 1818 og vil ikke skje igjen før i 2285.
første_pinsedag = første_påskedag + timedelta(days=49) #I de vestlige kirkene er den tidligst mulige datoen for pinsedag 10. mai, den senest mulige er 13. juni.
andre_pinsedag = første_påskedag + timedelta(days=50)

#første_påskedag_måned = første_påskedag.month
#print(ws_turnus_jan['A5'].font.color.rgb) #ikke slett

#Endre farge, dato og linjekant på siste rad i februar hvis det er skuddår
medium_border = Border(left=Side(style='medium'), 
            right=Side(style='medium'), 
            top=Side(style='medium'), 
            bottom=Side(style='medium'))

medium_og_tynn_border = Border(left=Side(style='thin'), 
            right=Side(style='medium'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin'))

medium_bunn = Border(left=Side(style='medium'), 
            bottom=Side(style='medium'))

medium_bunn1 = Border(left=Side(style='thin'), 
            bottom=Side(style='medium'))

test = Border(left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin'))

# Endre slutten av februar-turnusen hvis det er skuddår
if skuddår == True:
    # Slå sammen siste cellen etter 29. feb
    ws_turnus_feb.merge_cells(start_row=3, start_column=31, end_row=3, end_column=32)
    # Endre datoer på begynnelsen av mars
    ws_turnus_feb.cell(column=februar_skuddår+1,row=4, value=29)
    ws_turnus_feb.cell(column=februar_skuddår+2,row=4, value=1)
    ws_turnus_feb.cell(column=februar_skuddår+3,row=4, value=2)
    
    #Endre farge på dato-celle
    ws_turnus_feb['AD4'].fill = PatternFill(start_color='FFFFFF99', fill_type = 'solid')
    
    # Endre  font på dager og dato
    ws_turnus_feb['AD4'].font = Font(color = "FF000000", name= 'Calibri', size=12, bold=True)
    ws_turnus_feb['AD5'].font = Font(color = "FF00B050", name= 'Calibri', size=11, bold=True)
    
    # Endre font på vakter
    for rows in ws_turnus_feb.iter_rows(min_row=6, max_row=20, min_col=30, max_col=30):
        for cell in rows:
            cell.font = Font(color = "FF000000", name= 'Calibri', size=11, bold=True)
            
    # Endre bakgrunnsfarge på vakter til hvit
    for rows in ws_turnus_feb.iter_rows(min_row=5, max_row=20, min_col=30, max_col=30):
        for cell in rows:
            cell.fill = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type = "solid")
    
    # Fyll inn tekst MARS is siste celle og midstill
    ws_turnus_feb.cell(column=31,row=3, value="MARS")
    ws_turnus_feb.cell(column=31,row=3).alignment = Alignment(horizontal='center') 
    
    # Endre border på cellene (dette kan gjøres mye bedre...)
    for rows in ws_turnus_feb.iter_rows(min_row=4, max_row=20, min_col=30, max_col=30):
        for cell in rows:
            cell.border = medium_og_tynn_border
            
    for rows in ws_turnus_feb.iter_rows(min_row=4, max_row=20, min_col=29, max_col=29):
        for cell in rows:
            cell.border = test
    
    ws_turnus_feb.cell(row=4, column=30).border = medium_bunn1
    ws_turnus_feb.cell(row=5, column=30).border = medium_bunn1
    ws_turnus_feb.cell(row=11, column=30).border = medium_bunn1
    ws_turnus_feb.cell(row=13, column=30).border = medium_bunn1
    ws_turnus_feb.cell(row=20, column=30).border = medium_bunn1
    
    ws_turnus_feb.cell(row=4, column=29).border = medium_bunn1
    ws_turnus_feb.cell(row=5, column=29).border = medium_bunn1
    ws_turnus_feb.cell(row=11, column=29).border = medium_bunn1
    ws_turnus_feb.cell(row=13, column=29).border = medium_bunn1
    ws_turnus_feb.cell(row=20, column=29).border = medium_bunn1
    
    ws_turnus_feb.cell(row=3, column=31).border = medium_border
    ws_turnus_feb.cell(row=4, column=31).border = medium_bunn
    ws_turnus_feb.cell(row=5, column=31).border = medium_bunn
    ws_turnus_feb.cell(row=11, column=31).border = medium_bunn
    ws_turnus_feb.cell(row=13, column=31).border = medium_bunn
    ws_turnus_feb.cell(row=20, column=31).border = medium_bunn
    
else:
    ws_turnus_feb.merge_cells(start_row=3, start_column=30, end_row=3, end_column=32)
    ws_turnus_feb.cell(column=30,row=3, value="MARS")
    ws_turnus_feb.cell(column=30,row=3).alignment = Alignment(horizontal='center') 
    ws_turnus_feb.cell(column=30,row=4, value=1)
    ws_turnus_feb.cell(column=31,row=4, value=2)
    ws_turnus_feb.cell(column=32,row=4, value=3)

class Uker:
    def __init__(self, måned_tall, måned_navn, start_tur, slutt_tur, siste_celle):
        self.måned_tall = måned_tall
        self.måned_navn = måned_navn
        self.start_tur = start_tur
        self.slutt_tur = slutt_tur
        self.siste_celle = siste_celle
    
    # Få ukene for hver måned, maks 6 uker. En liste for hver uke,
    def fyll_inn_uker(self):
        ws_turnus = wb_turnus[self.måned_navn]
        
        # Fyll inn helger
        måned_nummer = self.måned_tall
        måned_slutt = self.siste_celle-1
        
        lørdag =pd.date_range(f'{år}-{måned_nummer}-01',f'{år}-{måned_nummer}-{måned_slutt}',freq='W-SAT')
        søndag =pd.date_range(f'{år}-{måned_nummer}-01',f'{år}-{måned_nummer}-{måned_slutt}',freq='W-SUN')
        #print(lørdag_januar[3].day)

        #søndager
        try:
            for rows in ws_turnus.iter_rows(min_row=4, max_row=20, min_col=søndag[0].day+1, max_col=søndag[0].day+1):
                for cell in rows:
                    cell.fill = PatternFill(start_color='FFFFFF99', end_color='FFFFFF99', fill_type = "solid")

            for rows in ws_turnus.iter_rows(min_row=4, max_row=20, min_col=søndag[1].day+1, max_col=søndag[1].day+1):
                for cell in rows:
                    cell.fill = PatternFill(start_color='FFFFFF99', end_color='FFFFFF99', fill_type = "solid")
                    
            for rows in ws_turnus.iter_rows(min_row=4, max_row=20, min_col=søndag[2].day+1, max_col=søndag[2].day+1):
                for cell in rows:
                    cell.fill = PatternFill(start_color='FFFFFF99', end_color='FFFFFF99', fill_type = "solid")
                    
            for rows in ws_turnus.iter_rows(min_row=4, max_row=20, min_col=søndag[3].day+1, max_col=søndag[3].day+1):
                for cell in rows:
                    cell.fill = PatternFill(start_color='FFFFFF99', end_color='FFFFFF99', fill_type = "solid")
                    
            for rows in ws_turnus.iter_rows(min_row=4, max_row=20, min_col=søndag[4].day+1, max_col=søndag[4].day+1):
                for cell in rows:
                    cell.fill = PatternFill(start_color='FFFFFF99', end_color='FFFFFF99', fill_type = "solid")
                    
            for rows in ws_turnus.iter_rows(min_row=4, max_row=20, min_col=søndag[5].day+1, max_col=søndag[5].day+1):
                for cell in rows:
                    cell.fill = PatternFill(start_color='FFFFFF99', end_color='FFFFFF99', fill_type = "solid")
        except IndexError:
            pass

        #lørdager
        try:
            for rows in ws_turnus.iter_rows(min_row=4, max_row=20, min_col=lørdag[0].day+1, max_col=lørdag[0].day+1):
                for cell in rows:
                    cell.fill = PatternFill(start_color='FFFFFF99', end_color='FFFFFF99', fill_type = "solid")

            for rows in ws_turnus.iter_rows(min_row=4, max_row=20, min_col=lørdag[1].day+1, max_col=lørdag[1].day+1):
                for cell in rows:
                    cell.fill = PatternFill(start_color='FFFFFF99', end_color='FFFFFF99', fill_type = "solid")
                    
            for rows in ws_turnus.iter_rows(min_row=4, max_row=20, min_col=lørdag[2].day+1, max_col=lørdag[2].day+1):
                for cell in rows:
                    cell.fill = PatternFill(start_color='FFFFFF99', end_color='FFFFFF99', fill_type = "solid")
                    
            for rows in ws_turnus.iter_rows(min_row=4, max_row=20, min_col=lørdag[3].day+1, max_col=lørdag[3].day+1):
                for cell in rows:
                    cell.fill = PatternFill(start_color='FFFFFF99', end_color='FFFFFF99', fill_type = "solid")
                    
            for rows in ws_turnus.iter_rows(min_row=4, max_row=20, min_col=lørdag[4].day+1, max_col=lørdag[4].day+1):
                for cell in rows:
                    cell.fill = PatternFill(start_color='FFFFFF99', end_color='FFFFFF99', fill_type = "solid")
                    
            for rows in ws_turnus.iter_rows(min_row=4, max_row=20, min_col=lørdag[5].day+1, max_col=lørdag[5].day+1):
                for cell in rows:
                    cell.fill = PatternFill(start_color='FFFFFF99', end_color='FFFFFF99', fill_type = "solid")
        except IndexError:
            pass
        
        # Hent ut uker pr' måned og filtrer ut 0-verdier i første og siste uke
        # Try i tilfelle det er ferre enn 6 ukenummer pr. måned
        try:
            uke1 = list(filter(None, calendar.monthcalendar(år, self.måned_tall)[0]))
            uke2 = list(filter(None, calendar.monthcalendar(år, self.måned_tall)[1]))
            uke3 = list(filter(None, calendar.monthcalendar(år, self.måned_tall)[2]))
            uke4 = list(filter(None, calendar.monthcalendar(år, self.måned_tall)[3]))
            uke5 = list(filter(None, calendar.monthcalendar(år, self.måned_tall)[4]))
            uke6 = list(filter(None, calendar.monthcalendar(år, self.måned_tall)[5]))
        except IndexError:
            pass

        # Hent siste dag i ukene i januar for å vite hvor cellene skal merge.
        #Try, i tilfelle alle uke-variablene ikke har blitt laget
        try:
            siste_ukedag_1 = uke1[-1]
            siste_ukedag_2 = uke2[-1]
            siste_ukedag_3 = uke3[-1]
            siste_ukedag_4 = uke4[-1]  
            siste_ukedag_5 = uke5[-1]
            siste_ukedag_6 = uke6[-1]
        except UnboundLocalError:
            pass
        
        # Merge celle for uker
        # Bruk try i tilfelle en uke kun bruker en celle
        try:
            ws_turnus.merge_cells(start_row=3, start_column=2, end_row=3, end_column=siste_ukedag_1+1)
            ws_turnus.merge_cells(start_row=3, start_column=siste_ukedag_1+2, end_row=3, end_column=siste_ukedag_2+1)
            ws_turnus.merge_cells(start_row=3, start_column=siste_ukedag_2+2, end_row=3, end_column=siste_ukedag_3+1)
            ws_turnus.merge_cells(start_row=3, start_column=siste_ukedag_3+2, end_row=3, end_column=siste_ukedag_4+1)
            ws_turnus.merge_cells(start_row=3, start_column=siste_ukedag_4+2, end_row=3, end_column=siste_ukedag_5+1)
            ws_turnus.merge_cells(start_row=3, start_column=siste_ukedag_5+2, end_row=3, end_column=siste_ukedag_6+1)
        except UnboundLocalError:
            pass

        try:
            liste_ukenummer = []
            liste_ukenummer.append(date(år, self.måned_tall, siste_ukedag_1).isocalendar()[1]) #isocalender dokuemntasjon
            liste_ukenummer.append(date(år, self.måned_tall, siste_ukedag_2).isocalendar()[1]) # bruk .week() istedenfor[1] i python 3.9
            liste_ukenummer.append(date(år, self.måned_tall, siste_ukedag_3).isocalendar()[1])
            liste_ukenummer.append(date(år, self.måned_tall, siste_ukedag_4).isocalendar()[1])
            liste_ukenummer.append(date(år, self.måned_tall, siste_ukedag_5).isocalendar()[1])
            liste_ukenummer.append(date(år, self.måned_tall, siste_ukedag_6).isocalendar()[1])
        except UnboundLocalError:
            pass

        #c = 2
        #for ukenummer in liste_ukenummer:
        #    ws_turnus.cell(row=3, column=c, value = ukenummer)
        #    c += siste_ukedag_1+2
        try:
            ws_turnus.cell(row=3, column=2, value=liste_ukenummer[0])
            ws_turnus.cell(row=3, column=siste_ukedag_1+2, value=liste_ukenummer[1])
            ws_turnus.cell(row=3, column=siste_ukedag_2+2, value=liste_ukenummer[2])
            ws_turnus.cell(row=3, column=siste_ukedag_3+2, value=liste_ukenummer[3])
            ws_turnus.cell(row=3, column=siste_ukedag_4+2, value=liste_ukenummer[4])
            ws_turnus.cell(row=3, column=siste_ukedag_5+2, value=liste_ukenummer[5])
        except IndexError:
            pass
        
        # Endre tekst på ukenummer etter merge
        tekst_uker = Font(bold=True) # Kan legge til flere ting
        for rows in ws_turnus.iter_rows(min_row=3, max_row=3, min_col=2, max_col=self.siste_celle):
            for cell in rows:
                cell.font = tekst_uker
        
        # Midstill tekst på ukenummer etter merge
        for rows in ws_turnus.iter_rows(min_row=3, max_row=3, min_col=2, max_col=self.siste_celle):
            for cell in rows:
                cell.alignment = Alignment(horizontal='center')      
        #ws_turnus['B5'].font = Font(bold=True)

        dato_idag = date.today().strftime('%d.%m.%Y')
        
        # Legg inn overskrift med år og måned
        ws_turnus.cell(column=1,row=1, value=self.måned_navn + ' ' + str(år))
        
        # Legg inn dato og signatur for når turnusen er oppdatert
        ws_turnus.cell(column=1,row=2, value="oppdatert " + dato_idag + " osh")
        
        # Fylle inn ukedager
        # Henter alle ukedageen for hver måned fra Pandas, oversetter til norsk med Locale
        dager_navn = pd.date_range(f'{år}-{måned_nummer}-01',f'{år}-{måned_nummer}-{måned_slutt}').day_name(locale ='no_NO.utf8').tolist()
        
        # Henter ut forbokstav til dager i liste dager_navn
        dager_bokstav = []
        for dager in dager_navn:
            dager_bokstav.append(dager[0])
            
        # Legger til forbokstavene til dagene i turnusen
        c_dager = 2
        for dager1 in dager_bokstav:
            ws_turnus.cell(row=5, column=c_dager).value = dager1
            c_dager += 1
        
        # Turnusvakter
        # hent vakter fra turnusmal, kjør forlopp for å få verdiene fra cellene.
        vakter_temp = ws_turnus_mal['B10':'AQ10']
        liste_vakter = []
        for cell in vakter_temp:
            for y in cell:
                liste_vakter.append(y.value)
        liste_vakter_år = liste_vakter * 10
        
        # Legg til vakter i turnus
        c = 2
        for vakter in liste_vakter_år[self.start_tur:self.slutt_tur]:
            ws_turnus.cell(row=9, column=c).value = vakter
            c += 1
        
        c1 = 2
        for vakter in liste_vakter_år[self.start_tur+7:self.slutt_tur+7]:
            ws_turnus.cell(row=8, column=c1).value = vakter
            c1 += 1
            
        c2 = 2
        for vakter in liste_vakter_år[self.start_tur+14:self.slutt_tur+14]:
            ws_turnus.cell(row=10, column=c2).value = vakter
            c2 += 1
            
        c3 = 2
        for vakter in liste_vakter_år[self.start_tur+21:self.slutt_tur+21]:
            ws_turnus.cell(row=7, column=c3).value = vakter
            c3 += 1
            
        c4 = 2
        for vakter in liste_vakter_år[self.start_tur+28:self.slutt_tur+28]:
            ws_turnus.cell(row=11, column=c4).value = vakter
            c4 += 1
            
        c5 = 2
        for vakter in liste_vakter_år[self.start_tur+35:self.slutt_tur+35]:
            ws_turnus.cell(row=6, column=c5).value = vakter
            c5 += 1
        
# Regner ut hvor turnusen starter hver måned
start_januar = 12 #liste på 42 vakter, trekk fra 13 for hvert år
slutt_januar = start_januar+31

start_februar = slutt_januar
slutt_februar = start_februar+februar_skuddår+ekstra_dager

start_mars = slutt_februar-ekstra_dager
slutt_mars = start_mars+31

start_april = slutt_mars
slutt_april = start_april+30+1

start_mai = slutt_april-1
slutt_mai = start_mai+31

start_juni = slutt_mai
slutt_juni = start_juni+30+1

start_juli = slutt_juni-1
slutt_juli = start_juli+31

start_august = slutt_juli
slutt_august = start_august+31

start_september = slutt_august
slutt_september = start_september+30+1

start_oktober = slutt_september-1
slutt_oktober = start_oktober+31

start_november = slutt_oktober
slutt_november = start_november+30+1

start_desember = slutt_november-1
slutt_desember = start_desember+31

# Definer objekt ( måned_tall, måned_navn, start_tur, slutt_tur, siste_celle)
måned_jan = Uker(1, 'Januar', start_januar, slutt_januar, 32)
måned_feb = Uker(2, 'Februar', start_februar, slutt_februar, februar_skuddår+1)
måned_mar = Uker(3, 'Mars', start_mars, slutt_mars, 32)
måned_apr = Uker(4, 'April', start_april, slutt_april, 31)
måned_mai = Uker(5, 'Mai', start_mai, slutt_mai, 32)
måned_jun = Uker(6, 'Juni', start_juni, slutt_juni, 31)
måned_jul = Uker(7, 'Juli', start_juli, slutt_juli, 32)
måned_aug = Uker(8, 'August', start_august, slutt_august, 32)
måned_sep = Uker(9, 'September', start_september, slutt_september, 31)
måned_okt = Uker(10, 'Oktober', start_oktober, slutt_oktober, 32)
måned_nov = Uker(11, 'November', start_november, slutt_november, 31)
måned_des = Uker(12, 'Desember', start_desember, slutt_desember, 32)

# Kjør funksjon for hver måned
måned_jan.fyll_inn_uker()
måned_feb.fyll_inn_uker()
måned_mar.fyll_inn_uker()
måned_apr.fyll_inn_uker()
måned_mai.fyll_inn_uker()
måned_jun.fyll_inn_uker()
måned_jul.fyll_inn_uker()
måned_aug.fyll_inn_uker()
måned_sep.fyll_inn_uker()
måned_okt.fyll_inn_uker()
måned_nov.fyll_inn_uker()
måned_des.fyll_inn_uker()

# Fyll inn bakgrunnsfarge på faste helligdager(må gjøres i tilefelle fargen til helgene dekker over helligdager eller flaggdager) 

# 1. januar nyttårsdag
for rows in ws_turnus_jan.iter_rows(min_row=6, max_row=20, min_col=2, max_col=2):
    for cell in rows:
        cell.fill = PatternFill(start_color='FFFF9999', end_color='FFFF9999', fill_type = "solid")
for rows in ws_turnus_jan.iter_rows(min_row=4, max_row=5, min_col=2, max_col=2):
    for cell in rows:
        cell.fill = PatternFill(start_color='FFB8CCE4', end_color='FFB8CCE4', fill_type = "solid")

#21. januar HKH Prinsesse Ingrid Alexandria f.2004
for rows in ws_turnus_jan.iter_rows(min_row=4, max_row=20, min_col=22, max_col=22):
    for cell in rows:
        cell.fill = PatternFill(start_color='FFB8CCE4', end_color='FFB8CCE4', fill_type = "solid")

# 6. februar: Samefolkets dag
for rows in ws_turnus_feb.iter_rows(min_row=4, max_row=20, min_col=7, max_col=7):
    for cell in rows:
        cell.fill = PatternFill(start_color='FFB8CCE4', end_color='FFB8CCE4', fill_type = "solid")
        
#21. februar: H.M.K. Haralds V.s fødselsdag
for rows in ws_turnus_feb.iter_rows(min_row=4, max_row=20, min_col=22, max_col=22):
    for cell in rows:
        cell.fill = PatternFill(start_color='FFB8CCE4', end_color='FFB8CCE4', fill_type = "solid")
        
#1. mai - Off. høytidsdag
for rows in ws_turnus_mai.iter_rows(min_row=6, max_row=20, min_col=2, max_col=2):
    for cell in rows:
        cell.fill = PatternFill(start_color='FFFF9999', end_color='FFFF9999', fill_type = "solid")
for rows in ws_turnus_mai.iter_rows(min_row=4, max_row=5, min_col=2, max_col=2):
    for cell in rows:
        cell.fill = PatternFill(start_color='FFB8CCE4', end_color='FFB8CCE4', fill_type = "solid")
        
#8. mai - Frigj./Veteran
for rows in ws_turnus_mai.iter_rows(min_row=4, max_row=20, min_col=9, max_col=9):
    for cell in rows:
        cell.fill = PatternFill(start_color='FFB8CCE4', end_color='FFB8CCE4', fill_type = "solid")

#17. mai - tas etter de variable helligdagene. I tilfelle pinsen treffer på 17. mai, så forsvinner ikke flaggdag
        
#7. juni Unionsoppløsningen (i år 2049 vil Unionsoppløsningen falle på samme dag som pinse)
for rows in ws_turnus_jun.iter_rows(min_row=4, max_row=20, min_col=8, max_col=8):
    for cell in rows:
        cell.fill = PatternFill(start_color='FFB8CCE4', end_color='FFB8CCE4', fill_type = "solid")
        
#4. juli - Dronning Sonja
for rows in ws_turnus_jul.iter_rows(min_row=4, max_row=20, min_col=5, max_col=5):
    for cell in rows:
        cell.fill = PatternFill(start_color='FFB8CCE4', end_color='FFB8CCE4', fill_type = "solid")
        
#20. juli - Kronprins Haakon
for rows in ws_turnus_jul.iter_rows(min_row=4, max_row=20, min_col=21, max_col=21):
    for cell in rows:
        cell.fill = PatternFill(start_color='FFB8CCE4', end_color='FFB8CCE4', fill_type = "solid")
        
#29. juli - Olsok
for rows in ws_turnus_jul.iter_rows(min_row=4, max_row=20, min_col=30, max_col=30):
    for cell in rows:
        cell.fill = PatternFill(start_color='FFB8CCE4', end_color='FFB8CCE4', fill_type = "solid")
        
#19. august - Kronprinsesse Mette-Marit
for rows in ws_turnus_aug.iter_rows(min_row=4, max_row=20, min_col=20, max_col=20):
    for cell in rows:
        cell.fill = PatternFill(start_color='FFB8CCE4', end_color='FFB8CCE4', fill_type = "solid")

# 24. desember - Julaften + 1. og 2. juledag
for rows in ws_turnus_des.iter_rows(min_row=4, max_row=20, min_col=25, max_col=27):
    for cell in rows:
        cell.fill = PatternFill(start_color='FFFF9999', end_color='FFFF9999', fill_type = "solid")
        
for rows in ws_turnus_des.iter_rows(min_row=4, max_row=5, min_col=26, max_col=26):
    for cell in rows:
        cell.fill = PatternFill(start_color='FFB8CCE4', end_color='FFB8CCE4', fill_type = "solid")
        
#31. desember - Nyttårsaften
for rows in ws_turnus_des.iter_rows(min_row=4, max_row=20, min_col=32, max_col=32):
    for cell in rows:
        cell.fill = PatternFill(start_color='FFFF9999', end_color='FFFF9999', fill_type = "solid")


# Sjekk når det er bevegelige helligdager og flaggdager og fyll inn bakgrunnsfarge
påskedager_april = []
påskedager_mars = []

# Skjærtorsdag
if skjærtorsdag.month == 3:
    påskedager_mars.append(str(skjærtorsdag.day)+". mars - Skjærtorsdag")
    for rows in ws_turnus_mar.iter_rows(min_row=4, max_row=20, min_col=skjærtorsdag.day+1, max_col=skjærtorsdag.day+1):
        for cell in rows:
            cell.fill = PatternFill(start_color='FFFF9999', end_color='FFFF9999', fill_type = "solid")

if skjærtorsdag.month == 4:
    påskedager_april.append(str(skjærtorsdag.day)+". mars - Skjærtorsdag")
    for rows in ws_turnus_apr.iter_rows(min_row=4, max_row=20, min_col=skjærtorsdag.day+1, max_col=skjærtorsdag.day+1):
        for cell in rows:
            cell.fill = PatternFill(start_color='FFFF9999', end_color='FFFF9999', fill_type = "solid")

# Langfredag
if langfredag.month == 3:
    påskedager_mars.append(str(langfredag.day)+". mars - Langfredag")
    for rows in ws_turnus_mar.iter_rows(min_row=4, max_row=20, min_col=langfredag.day+1, max_col=langfredag.day+1):
        for cell in rows:
            cell.fill = PatternFill(start_color='FFFF9999', end_color='FFFF9999', fill_type = "solid")

if langfredag.month == 4:
    påskedager_april.append(str(langfredag.day)+". april - Langfredag")
    for rows in ws_turnus_apr.iter_rows(min_row=4, max_row=20, min_col=langfredag.day+1, max_col=langfredag.day+1):
        for cell in rows:
            cell.fill = PatternFill(start_color='FFFF9999', end_color='FFFF9999', fill_type = "solid")

# Påskeaften
if påskeaften.month == 3:
    påskedager_mars.append(str(påskeaften.day)+". mars - Påskeaften")
    for rows in ws_turnus_mar.iter_rows(min_row=4, max_row=20, min_col=påskeaften.day+1, max_col=påskeaften.day+1):
        for cell in rows:
            cell.fill = PatternFill(start_color='FFFF9999', end_color='FFFF9999', fill_type = "solid")

if påskeaften.month == 4:
    påskedager_april.append(str(påskeaften.day)+". april - Påskeaften")
    for rows in ws_turnus_apr.iter_rows(min_row=4, max_row=20, min_col=påskeaften.day+1, max_col=påskeaften.day+1):
        for cell in rows:
            cell.fill = PatternFill(start_color='FFFF9999', end_color='FFFF9999', fill_type = "solid")

# Første påskedag
if første_påskedag.month == 3:
    påskedager_mars.append(str(første_påskedag.day)+". mars - Første påskedag")
    for rows in ws_turnus_mar.iter_rows(min_row=4, max_row=5, min_col=første_påskedag.day+1, max_col=første_påskedag.day+1):
        for cell in rows:
            cell.fill = PatternFill(start_color='FFB8CCE4', end_color='FFB8CCE4', fill_type = "solid")
    for rows in ws_turnus_mar.iter_rows(min_row=4, max_row=20, min_col=første_påskedag.day+1, max_col=første_påskedag.day+1):
        for cell in rows:
            cell.fill = PatternFill(start_color='FFFF9999', end_color='FFFF9999', fill_type = "solid")

if første_påskedag.month == 4:
    påskedager_april.append(str(første_påskedag.day)+". april - Første påskedag")
    for rows in ws_turnus_apr.iter_rows(min_row=4, max_row=5, min_col=første_påskedag.day+1, max_col=første_påskedag.day+1):
        for cell in rows:
            cell.fill = PatternFill(start_color='FFB8CCE4', end_color='FFB8CCE4', fill_type = "solid")
    for rows in ws_turnus_apr.iter_rows(min_row=6, max_row=20, min_col=første_påskedag.day+1, max_col=første_påskedag.day+1):
        for cell in rows:
            cell.fill = PatternFill(start_color='FFFF9999', end_color='FFFF9999', fill_type = "solid")

# Andre påskedag
if andre_påskedag.month == 3:
    påskedager_mars.append(str(andre_påskedag.day)+". mars - Andre påskedag")
    for rows in ws_turnus_mar.iter_rows(min_row=4, max_row=20, min_col=andre_påskedag.day+1, max_col=andre_påskedag.day+1):
        for cell in rows:
            cell.fill = PatternFill(start_color='FFFF9999', end_color='FFFF9999', fill_type = "solid")

if andre_påskedag.month == 4:
    påskedager_april.append(str(andre_påskedag.day)+". april - Andre påskedag")
    for rows in ws_turnus_apr.iter_rows(min_row=4, max_row=20, min_col=andre_påskedag.day+1, max_col=andre_påskedag.day+1):
        for cell in rows:
            cell.fill = PatternFill(start_color='FFFF9999', end_color='FFFF9999', fill_type = "solid")

#Fyll inn navn på høytidsdager nederst på turnusen
r = 22
for påskedager in påskedager_mars:
    ws_turnus_mar.cell(row=r, column=2).value = påskedager
    r += 1

r1 = 22
for påskedager in påskedager_april:
    ws_turnus_apr.cell(row=r1, column=2).value = påskedager
    r1 += 1

#Fyll inn bakgrunnsfarge på høytidsdager nederst på turnusen
l_mars = len(påskedager_mars)
l_april = len(påskedager_april)

for rows in ws_turnus_mar.iter_rows(min_row=22, max_row=21+l_mars, min_col=2, max_col=2):
    for cell in rows:
        cell.fill = PatternFill(start_color='FFFF9999', end_color='FFFF9999', fill_type = "solid")

for rows in ws_turnus_apr.iter_rows(min_row=22, max_row=21+l_april, min_col=2, max_col=2):
    for cell in rows:
        cell.fill = PatternFill(start_color='FFFF9999', end_color='FFFF9999', fill_type = "solid")

#Lag kantlinjen på høytidsdager nederst på turnusen (kan gjøres bedre)
def set_border(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    medium = Side(border_style="medium", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=medium, right=medium, bottom=thin)

set_border(ws_turnus_mar, 'B22:K'+str(21+l_mars))
set_border(ws_turnus_apr, 'B22:K'+str(21+l_april))

nederste_celle = Border(left=Side(style='medium'), 
            right=Side(style='medium'), 
            bottom=Side(style='medium'))

if l_mars >=1:
    for rows in ws_turnus_mar.iter_rows(min_row=21+l_mars, max_row=21+l_mars, min_col=2, max_col=11):
        for cell in rows:
            cell.border = nederste_celle

if l_april >=1:
    for rows in ws_turnus_apr.iter_rows(min_row=21+l_april, max_row=21+l_april, min_col=2, max_col=11):
        for cell in rows:
            cell.border = nederste_celle


høytidsdager_mai = [] #Liste for å sortere høytidsdager i mai etter dato, siden noen er variable og andre ikke
høytidsdager_juni = []

# Kristi himmelfartsdag
if himmelspretten.month == 5:
    høytidsdager_mai.append(str(himmelspretten.day)+". mai - Kristi himmelfartsdag")
    for rows in ws_turnus_mai.iter_rows(min_row=4, max_row=20, min_col=himmelspretten.day+1, max_col=himmelspretten.day+1):
        for cell in rows:
            cell.fill = PatternFill(start_color='FFFF9999', end_color='FFFF9999', fill_type = "solid")
if himmelspretten.month == 6:
    høytidsdager_juni.append(str(himmelspretten.day)+". juni - Kristi himmelfartsdag")
    for rows in ws_turnus_jun.iter_rows(min_row=4, max_row=20, min_col=himmelspretten.day+1, max_col=himmelspretten.day+1):
        for cell in rows:
            cell.fill = PatternFill(start_color='FFFF9999', end_color='FFFF9999', fill_type = "solid")

#8. mai - Frigj./Veteran
høytidsdager_mai.append("8. mai - Frigj./Veteran")

# Første pinsedag
if første_pinsedag.month == 5:
    høytidsdager_mai.append(str(første_pinsedag.day)+". mai - 1. pinsedag")
    for rows in ws_turnus_mai.iter_rows(min_row=4, max_row=5, min_col=første_pinsedag.day+1, max_col=første_pinsedag.day+1):
        for cell in rows:
            cell.fill = PatternFill(start_color='FFB8CCE4', end_color='FFB8CCE4', fill_type = "solid")
    for rows in ws_turnus_mai.iter_rows(min_row=6, max_row=20, min_col=første_pinsedag.day+1, max_col=første_pinsedag.day+1):
        for cell in rows:
            cell.fill = PatternFill(start_color='FFFF9999', end_color='FFFF9999', fill_type = "solid")
if første_pinsedag.month == 6:
    høytidsdager_juni.append(str(første_pinsedag.day)+". juni - 1. pinsedag")
    for rows in ws_turnus_jun.iter_rows(min_row=4, max_row=5, min_col=første_pinsedag.day+1, max_col=første_pinsedag.day+1):
        for cell in rows:
            cell.fill = PatternFill(start_color='FFB8CCE4', end_color='FFB8CCE4', fill_type = "solid")
    for rows in ws_turnus_jun.iter_rows(min_row=6, max_row=20, min_col=første_pinsedag.day+1, max_col=første_pinsedag.day+1):
        for cell in rows:
            cell.fill = PatternFill(start_color='FFFF9999', end_color='FFFF9999', fill_type = "solid")

# Andre pinsedag
if andre_pinsedag.month == 5:
    høytidsdager_mai.append(str(andre_pinsedag.day)+". mai - 2. pinsedag")
    for rows in ws_turnus_mai.iter_rows(min_row=4, max_row=20, min_col=andre_pinsedag.day+1, max_col=andre_pinsedag.day+1):
        for cell in rows:
            cell.fill = PatternFill(start_color='FFFF9999', end_color='FFFF9999', fill_type = "solid")

if andre_pinsedag.month == 6:
    høytidsdager_juni.append(str(andre_pinsedag.day)+". juni - 2. pinsedag")
    ws_turnus_jun['B24'].fill = PatternFill(start_color='FFFF9999', fill_type = 'solid')
    for rows in ws_turnus_jun.iter_rows(min_row=4, max_row=20, min_col=andre_pinsedag.day+1, max_col=andre_pinsedag.day+1):
        for cell in rows:
            cell.fill = PatternFill(start_color='FFFF9999', end_color='FFFF9999', fill_type = "solid")

#7. juni Unionsoppløsningen
høytidsdager_juni.append("7. juni Unionsoppløsningen")

# 17. mai - Grunnlovsdag
høytidsdager_mai.append("17. mai - Grunnlovsdagen")
for rows in ws_turnus_mai.iter_rows(min_row=6, max_row=20, min_col=18, max_col=18):
    for cell in rows:
        cell.fill = PatternFill(start_color='FFFF9999', end_color='FFFF9999', fill_type = "solid")
for rows in ws_turnus_mai.iter_rows(min_row=4, max_row=5, min_col=18, max_col=18):
    for cell in rows:
        cell.fill = PatternFill(start_color='FFB8CCE4', end_color='FFB8CCE4', fill_type = "solid")

#Sorter liste høytidsdager_mai
sorterte_høytidsdager_mai = sorted(høytidsdager_mai, key=lambda x: int(x.split('.')[0]))
sorterte_høytidsdager_juni = sorted(høytidsdager_juni, key=lambda x: int(x.split('.')[0]))

#Fyller inn navn på høydtidsdager nederst i turnusen 
r3 = 23
for høytidsdager in sorterte_høytidsdager_mai:
    ws_turnus_mai.cell(row=r3, column=2).value = høytidsdager
    r3 += 1
    
r4 = 22
for høytidsdager in sorterte_høytidsdager_juni:
    ws_turnus_jun.cell(row=r4, column=2).value = høytidsdager
    r4 += 1

# finne lngde på lister
l_mai = len(høytidsdager_mai)
l_juni = len(høytidsdager_juni)

for rows in ws_turnus_mai.iter_rows(min_row=23, max_row=22+l_mai, min_col=2, max_col=2):
    for cell in rows:
        cell.fill = PatternFill(start_color='FFFF9999', end_color='FFFF9999', fill_type = "solid")

for rows in ws_turnus_jun.iter_rows(min_row=22, max_row=21+l_juni, min_col=2, max_col=2):
    for cell in rows:
        cell.fill = PatternFill(start_color='FFFF9999', end_color='FFFF9999', fill_type = "solid")

#Kantlinjehelvete mai/juni
def set_border(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    medium = Side(border_style="medium", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=medium, right=medium, bottom=thin)

set_border(ws_turnus_mai, 'B22:K'+str(22+l_mai))
set_border(ws_turnus_jun, 'B22:K'+str(21+l_juni))

if l_mai >=1:
    for rows in ws_turnus_mai.iter_rows(min_row=22+l_mai, max_row=22+l_mai, min_col=2, max_col=11):
        for cell in rows:
            cell.border = nederste_celle

if l_juni >=1:
    for rows in ws_turnus_jun.iter_rows(min_row=21+l_juni, max_row=21+l_juni, min_col=2, max_col=11):
        for cell in rows:
            cell.border = nederste_celle

# Endre farge flaggdager nederst på turnus
index_fri = sorterte_høytidsdager_mai.index('8. mai - Frigj./Veteran') #henter nummer i liste
for rows in ws_turnus_mai.iter_rows(min_row=22+index_fri+1, max_row=22+index_fri+1, min_col=2, max_col=2):
    for cell in rows:
        cell.fill = PatternFill(start_color='FFB8CCE4', end_color='FFB8CCE4', fill_type = "solid")
        
index_uni = sorterte_høytidsdager_juni.index('7. juni Unionsoppløsningen')
for rows in ws_turnus_jun.iter_rows(min_row=21+index_uni+1, max_row=21+index_uni+1, min_col=2, max_col=2):
    for cell in rows:
        cell.fill = PatternFill(start_color='FFB8CCE4', end_color='FFB8CCE4', fill_type = "solid")

# Lagre ny turnus
wb_turnus.save('Turnus_'+str(år)+'.xlsx')

# Lukk fil
wb_turnus.close()

print("ferdig")